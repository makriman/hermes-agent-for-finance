"""Excel export — the PRD's 'augment Excel, don't compete' bridge.

Produces an .xlsx with three sheets:
  Forecast   line items x period buckets, opening/closing balance rows
  LineItems  the editable config (method, params, window, source)
  Actuals    last months by category (the data behind the forecast)
"""
from __future__ import annotations

import json
from collections import defaultdict
from datetime import date

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font

from . import taxonomy as tax

# LineItems sheet contract (consumed by engine/xlsx_import.py):
#   row 1  instruction banner (not data)
#   row 2  header row below
#   row 3+ one line item per row; ID is the stable store id
LINEITEMS_SHEET = "LineItems"
LINEITEMS_HEADERS = ["ID", "Name", "Category", "Method", "Params", "Start", "End",
                     "Source", "Locked", "Note"]
LINEITEMS_INSTRUCTIONS = (
    "HOW TO EDIT: you may change Name, Category, Method, Params (JSON), Start, End "
    "and Note. To add an item, add a row and leave ID blank. To end an item, set its "
    "End date. To remove an item, delete its row (it is deactivated, never deleted). "
    "Do not edit the ID column."
)


def _autofit(ws):
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 46)


def export_xlsx(path: str, result: dict, items: list[dict],
                enriched: list, months_back: int = 4) -> str:
    wb = Workbook()
    bold = Font(bold=True)

    # --- Forecast grid --------------------------------------------------------
    ws = wb.active
    ws.title = "Forecast"
    buckets = result["buckets"]
    heads = ["Line item", "Category"] + [f"{b['from'][5:]}→{b['to'][5:]}" for b in buckets] + ["Total"]
    ws.append(heads)
    for c_ in ws[1]:
        c_.font = bold

    # occurrence amounts bucketed per line label
    grid: dict[str, dict[int, float]] = defaultdict(lambda: defaultdict(float))
    label_cat: dict[str, str] = {}
    for o in result["occurrences"]:
        for i, b in enumerate(buckets):
            if b["from"] <= o.expected_date <= b["to"]:
                grid[o.label][i] += o.amount
                label_cat[o.label] = o.category
                break
    for label in sorted(grid, key=lambda l: (label_cat[l], l)):
        row = [label, tax.label(label_cat[label])]
        row += [round(grid[label].get(i, 0.0), 2) or None for i in range(len(buckets))]
        row.append(round(sum(grid[label].values()), 2))
        ws.append(row)

    ws.append([])
    net = ["Net movement", ""] + [b["net"] for b in buckets] + [result["projected_net"]]
    close = ["Closing balance", ""] + [b["close"] for b in buckets] + [result["close"]]
    ws.append(["Opening balance", "", result["opening_balance"]])
    ws.append(net)
    ws.append(close)
    for r in (ws.max_row - 2, ws.max_row - 1, ws.max_row):
        for c_ in ws[r]:
            c_.font = bold
    _autofit(ws)

    # --- LineItems config -----------------------------------------------------
    ws2 = wb.create_sheet(LINEITEMS_SHEET)
    ws2.append(LINEITEMS_HEADERS)
    for c_ in ws2[1]:
        c_.font = bold
    for it in items:
        ws2.append([it["id"], it["name"], it["category"], it["method"],
                    json.dumps(it["params"]), it["start_date"], it["end_date"],
                    it["source"], bool(it["locked"]), None])
    _autofit(ws2)
    # instruction banner goes in AFTER autofit so its length doesn't skew col A
    ws2.insert_rows(1)
    tip = ws2.cell(row=1, column=1, value=LINEITEMS_INSTRUCTIONS)
    tip.font = Font(italic=True, color="FF808080")
    id_head = ws2.cell(row=2, column=1)
    id_head.comment = Comment(
        "Stable store id — do not edit. Leave blank on new rows.", "cashew")

    # --- Actuals by category ---------------------------------------------------
    ws3 = wb.create_sheet("Actuals")
    months = sorted({t.date.strftime("%Y-%m") for t in enriched})[-months_back:]
    agg: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for t in enriched:
        ym = t.date.strftime("%Y-%m")
        if ym in months and not tax.is_transfer(t.category):
            agg[t.category][ym] += t.amount
    ws3.append(["Category"] + months)
    for c_ in ws3[1]:
        c_.font = bold
    for cat in sorted(agg):
        ws3.append([tax.label(cat)] + [round(agg[cat].get(m, 0.0), 2) for m in months])
    _autofit(ws3)

    wb.save(path)
    return path
