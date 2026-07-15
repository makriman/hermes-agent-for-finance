"""Engine v3 tests — editable line-item config, methods, compute, scenarios.

Needs the mock server on :8900 (skipped otherwise). Uses a throwaway DB.
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

import pytest

from engine import compute as cp
from engine import config_sync, methods as mt
from engine.client import Clients
from engine.store import Store

BASE = "http://127.0.0.1:8900"


@pytest.fixture(scope="module")
def c():
    cl = Clients(BASE)
    try:
        cl.set_org("jam-scn-1")
        cl.reset()
    except Exception:
        pytest.skip("mock server not running on :8900")
    return cl


@pytest.fixture()
def store(tmp_path):
    return Store(tmp_path / "t.db")


@pytest.fixture()
def synced(c, store):
    c.set_org("jam-scn-1")
    c.reset()
    cfg = c.config()
    config_sync.sync(c, store, "jam-scn-1", cfg["scenario_month"], cfg["anchor"])
    return store


# --- methods (pure) -------------------------------------------------------------

def test_recurring_fixed_with_growth():
    item = {"method": "recurring_fixed", "params": {"amount": -1000, "day_of_month": 10,
                                                    "growth_pct": 10}}
    occ = mt.project(item, ["2026-07", "2026-08"])
    assert [a for _, a in occ] == [-1000.0, -1100.0]
    assert occ[0][0] == date(2026, 7, 10)


def test_one_off_only_in_window():
    item = {"method": "one_off", "params": {"amount": -5000, "date": "2026-08-05"}}
    assert mt.project(item, ["2026-07"]) == []
    assert mt.project(item, ["2026-07", "2026-08"])[0][1] == -5000.0


def test_piecewise_start_end():
    item = {"method": "recurring_fixed", "params": {"amount": -4000, "day_of_month": 28},
            "start_date": "2026-09-01", "end_date": None}
    occ = mt.project(item, ["2026-07", "2026-08", "2026-09"])
    assert len(occ) == 1 and occ[0][0].month == 9


def test_trend_projection():
    item = {"method": "recurring_variable",
            "params": {"observations": [100, 200, 300], "lookback_months": 3,
                       "day_of_month": 1}}
    occ = mt.project(item, ["2026-07"])
    assert occ[0][1] == 400.0                     # perfect +100/mo trend continues


def test_linked_percentage():
    item = {"method": "linked", "params": {"pct": -30, "target": "revenue",
                                           "day_of_month": 15}}
    occ = mt.project_linked(item, ["2026-07"], {"2026-07": 10000.0})
    assert occ[0][1] == -3000.0


# --- config store ------------------------------------------------------------------

def test_item_crud_and_undo(store):
    iid = store.upsert_item("jam-scn-1", "Test item", "payroll", "recurring_fixed",
                            {"amount": -1000, "day_of_month": 5}, source="owner",
                            locked=1)
    assert store.get_item("jam-scn-1", iid)["params"]["amount"] == -1000
    store.upsert_item("jam-scn-1", "Test item", "payroll", "recurring_fixed",
                      {"amount": -2000, "day_of_month": 5}, source="owner", locked=1)
    assert store.get_item("jam-scn-1", iid)["params"]["amount"] == -2000
    msg = store.undo_last("jam-scn-1")
    assert "Reverted" in msg
    assert store.get_item("jam-scn-1", iid)["params"]["amount"] == -1000


def test_sync_never_touches_locked(synced, c):
    org = "jam-scn-1"
    items = synced.items(org)
    victim = items[0]["name"]
    synced.upsert_item(org, victim, items[0]["category"], "recurring_fixed",
                       {"amount": -99999, "day_of_month": 1}, source="owner", locked=1)
    cfg = c.config()
    rep = config_sync.sync(c, synced, org, cfg["scenario_month"], cfg["anchor"])
    assert victim in rep["skipped_locked"]
    assert synced.get_item(org, victim)["params"]["amount"] == -99999


def test_sync_seeds_config(synced):
    items = synced.items("jam-scn-1")
    assert len(items) >= 15
    assert all(it["method"] in ("recurring_fixed", "recurring_variable") for it in items)


# --- compute ---------------------------------------------------------------------------

def test_compute_multi_month_with_horizon_vat(c, synced):
    r = cp.compute(c, synced, "jam-scn-1", params={"horizon_months": 3})
    assert len(r["months"]) == 3
    vat_occ = [o for o in r["occurrences"] if o.source == "vat"]
    # v5 projects the UK cadence: payment due period-end + 1 month + 7 days.
    # Last paid 2026-05-11 -> current period ends 2026-07-31 -> due 2026-09-07.
    assert vat_occ and vat_occ[0].expected_date == "2026-09-07"
    assert r["vat"]["next_due"] == "2026-09-07"
    assert r["verdict"][0] in ("🔴", "🟡", "🟢")
    assert r["buckets"][-1]["close"] == r["close"]


def test_compute_scenario_overlay(c, synced):
    base = cp.compute(c, synced, "jam-scn-1")
    alt = cp.compute(c, synced, "jam-scn-1", scenario={"scale": {"revenue": 0.5}})
    assert alt["projected_net"] < base["projected_net"]


def test_compute_piecewise_new_hire(c, synced):
    synced.upsert_item("jam-scn-1", "New hire", "payroll", "recurring_fixed",
                       {"amount": -4000, "day_of_month": 28},
                       start_date="2026-09-01", source="owner", locked=1)
    r = cp.compute(c, synced, "jam-scn-1", params={"horizon_months": 3})
    hires = [o for o in r["occurrences"] if o.label == "New hire"]
    assert len(hires) == 1 and hires[0].expected_date.startswith("2026-09")


def test_compute_future_only_midmonth(c, synced):
    c.set_date("2026-07-15")
    r = cp.compute(c, synced, "jam-scn-1", params={"horizon_months": 1})
    assert all(o.expected_date > "2026-07-15" for o in r["occurrences"])
    c.reset()


def test_freeze_month_consistent(c, synced):
    fv = cp.freeze_month(c, synced, "jam-scn-1")
    assert fv.month == "2026-07"
    assert round(fv.opening_balance + fv.projected_net, 2) == fv.forecast_close
    assert fv.occurrences


# --- export ---------------------------------------------------------------------------

def test_export_xlsx(c, synced, tmp_path):
    from engine.actuals import build_enriched
    from engine.export import export_xlsx
    r = cp.compute(c, synced, "jam-scn-1")
    enriched, _, _, _ = build_enriched(c, c.config()["now"])
    out = export_xlsx(str(tmp_path / "f.xlsx"), r, synced.items("jam-scn-1"), enriched)
    from openpyxl import load_workbook
    wb = load_workbook(out)
    assert {"Forecast", "LineItems", "Actuals"} <= set(wb.sheetnames)
    assert wb["Forecast"].max_row > 5
