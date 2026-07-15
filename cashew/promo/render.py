"""Cashew v4 promo v2 — Allica-branded, 2.5x pacing, capabilities ending.
1080x1920 @ 30fps. Real bot outputs. Frames -> promo/frames/f%05d.png
"""
from __future__ import annotations

import math
import re
from pathlib import Path

from PIL import Image, ImageDraw, ImageFont

ROOT = Path(__file__).resolve().parent
FRAMES = ROOT / "frames"
FRAMES.mkdir(exist_ok=True)

W, H = 1080, 1920
FPS = 30

# ------------------------------------------------ theme (Telegram + Allica)
BG = "#0E1621"
HEADER = "#17212B"
IN_BUBBLE = "#182533"
OUT_BUBBLE = "#2B5278"
TEXT = "#F5F5F5"
SUBTLE = "#7D8E98"
ACCENT = "#5EB5F7"
GREEN = "#4FBF67"
RED = "#E85C5C"
AMBER = "#F5C542"
CHIP = "#1C2A38"
NAVY = "#1B2C5E"          # Allica navy
SLATE = "#7E8AA0"         # Allica byline grey
PAPER = "#FCFCFD"

DJ = "/usr/share/fonts/truetype/dejavu"
F = ImageFont.truetype(f"{DJ}/DejaVuSans.ttf", 34)
FB = ImageFont.truetype(f"{DJ}/DejaVuSans-Bold.ttf", 34)
FSM = ImageFont.truetype(f"{DJ}/DejaVuSans.ttf", 26)
FH = ImageFont.truetype(f"{DJ}/DejaVuSans-Bold.ttf", 40)
FLOGO_S = ImageFont.truetype(f"{DJ}/DejaVuSans-Bold.ttf", 54)
FMED = ImageFont.truetype(f"{DJ}/DejaVuSans.ttf", 44)
FMEDB = ImageFont.truetype(f"{DJ}/DejaVuSans-Bold.ttf", 46)
FPILL = ImageFont.truetype(f"{DJ}/DejaVuSans.ttf", 40)

LINE_H = 48
PAD = 26
BUBBLE_MAX = 880

ICONS = {"🟡": ("dot", AMBER), "🔴": ("dot", RED), "🟢": ("dot", GREEN),
         "✅": ("check", GREEN), "❌": ("cross", RED), "⏳": ("dot", "#E8A33D"),
         "❗": ("bang", RED), "💡": ("bulb", AMBER), "⚠️": ("tri", AMBER),
         "⚠": ("tri", AMBER), "📊": ("chart", GREEN)}
ICON_RE = re.compile("(" + "|".join(re.escape(k) for k in
                                    sorted(ICONS, key=len, reverse=True)) + ")")


def draw_icon(d, kind, color, x, y, s=34):
    cy = y + s // 2 + 3
    if kind == "dot":
        d.ellipse([x + 4, cy - 14, x + 32, cy + 14], fill=color)
    elif kind == "check":
        d.line([(x + 6, cy), (x + 15, cy + 10), (x + 32, cy - 12)], fill=color, width=6)
    elif kind == "cross":
        d.line([(x + 8, cy - 12), (x + 30, cy + 12)], fill=color, width=6)
        d.line([(x + 30, cy - 12), (x + 8, cy + 12)], fill=color, width=6)
    elif kind == "bang":
        d.line([(x + 19, cy - 14), (x + 19, cy + 6)], fill=color, width=7)
        d.ellipse([x + 15, cy + 11, x + 23, cy + 19], fill=color)
    elif kind == "bulb":
        d.ellipse([x + 8, cy - 15, x + 30, cy + 7], outline=color, width=4)
        d.line([(x + 15, cy + 10), (x + 23, cy + 10)], fill=color, width=4)
        d.line([(x + 16, cy + 16), (x + 22, cy + 16)], fill=color, width=4)
    elif kind == "tri":
        d.polygon([(x + 19, cy - 15), (x + 4, cy + 13), (x + 34, cy + 13)],
                  outline=color, width=4)
        d.line([(x + 19, cy - 4), (x + 19, cy + 4)], fill=color, width=4)
    elif kind == "chart":
        for i, h_ in enumerate((10, 20, 28)):
            d.rectangle([x + 6 + i * 10, cy + 14 - h_, x + 13 + i * 10, cy + 14], fill=color)
    return 40


def tokenize(line):
    out = []
    for part in ICON_RE.split(line):
        if not part:
            continue
        if part in ICONS:
            out.append(("icon", ICONS[part]))
            continue
        for seg in re.split(r"(\*[^*]+\*|_[^_]+_)", part):
            if not seg:
                continue
            if seg.startswith("*") and seg.endswith("*") and len(seg) > 2:
                out.append(("bold", seg[1:-1]))
            elif seg.startswith("_") and seg.endswith("_") and len(seg) > 2:
                out.append(("text", seg[1:-1]))
            else:
                out.append(("text", seg))
    return out


def seg_w(d, kind, payload):
    if kind == "icon":
        return 40
    return d.textlength(payload, font=FB if kind == "bold" else F)


def wrap_line(d, line, maxw):
    toks = tokenize(line)
    rows, cur, curw = [], [], 0
    for kind, payload in toks:
        if kind == "icon":
            if curw + 40 > maxw and cur:
                rows.append(cur)
                cur, curw = [], 0
            cur.append((kind, payload))
            curw += 40
            continue
        for word in re.split(r"(\s+)", payload):
            if not word:
                continue
            wlen = seg_w(d, kind, word)
            if curw + wlen > maxw and cur and word.strip():
                rows.append(cur)
                cur, curw = [], 0
                if not word.strip():
                    continue
            cur.append((kind, word))
            curw += wlen
    if cur:
        rows.append(cur)
    return rows or [[]]


def make_bubble(text, outgoing, time_s="09:41"):
    probe = ImageDraw.Draw(Image.new("RGB", (10, 10)))
    all_rows = []
    for ln in text.split("\n"):
        indent = 40 if ln.startswith("  ") else 0
        for r in wrap_line(probe, ln.strip() or " ", BUBBLE_MAX - 2 * PAD - indent):
            all_rows.append((indent, r))
    widest = max((ind + sum(seg_w(probe, k, p) for k, p in row)
                  for ind, row in all_rows), default=100)
    bw = int(min(BUBBLE_MAX, max(widest + 2 * PAD, 220)))
    bh = len(all_rows) * LINE_H + 2 * PAD + 20
    img = Image.new("RGBA", (bw, bh), (0, 0, 0, 0))
    d = ImageDraw.Draw(img)
    d.rounded_rectangle([0, 0, bw - 1, bh - 1], radius=26,
                        fill=OUT_BUBBLE if outgoing else IN_BUBBLE)
    y = PAD
    for indent, row in all_rows:
        x = PAD + indent
        for kind, payload in row:
            if kind == "icon":
                x += draw_icon(d, payload[0], payload[1], x, y)
            else:
                font = FB if kind == "bold" else F
                d.text((x, y), payload, font=font, fill=TEXT)
                x += d.textlength(payload, font=font)
        y += LINE_H
    stamp = time_s + ("  ✓✓" if outgoing else "")
    d.text((bw - PAD - probe.textlength(stamp, font=FSM), bh - 40),
           stamp, font=FSM, fill="#9AC0DF" if outgoing else SUBTLE)
    return img


def make_doc_bubble(filename, size, caption, time_s="09:42"):
    bw = 700
    probe = ImageDraw.Draw(Image.new("RGB", (10, 10)))
    cap_rows = wrap_line(probe, caption, bw - 2 * PAD)
    bh = 150 + len(cap_rows) * LINE_H + PAD + 26
    img = Image.new("RGBA", (bw, bh), (0, 0, 0, 0))
    d = ImageDraw.Draw(img)
    d.rounded_rectangle([0, 0, bw - 1, bh - 1], radius=26, fill=IN_BUBBLE)
    d.rounded_rectangle([PAD, PAD, PAD + 96, PAD + 96], radius=18, fill="#1D6F42")
    d.text((PAD + 22, PAD + 26), "X",
           font=ImageFont.truetype(f"{DJ}/DejaVuSans-Bold.ttf", 52), fill="white")
    for gy in (PAD + 20, PAD + 46, PAD + 72):
        d.line([(PAD + 62, gy), (PAD + 84, gy)], fill="white", width=4)
    d.text((PAD + 120, PAD + 8), filename, font=FB, fill=TEXT)
    d.text((PAD + 120, PAD + 56), size, font=FSM, fill=SUBTLE)
    y = 150
    for row in cap_rows:
        x = PAD
        for kind, payload in row:
            if kind == "icon":
                x += draw_icon(d, payload[0], payload[1], x, y)
            else:
                font = FB if kind == "bold" else F
                d.text((x, y), payload, font=font, fill=TEXT)
                x += d.textlength(payload, font=font)
        y += LINE_H
    d.text((bw - PAD - probe.textlength(time_s, font=FSM), bh - 40),
           time_s, font=FSM, fill=SUBTLE)
    return img


def make_chip(label):
    probe = ImageDraw.Draw(Image.new("RGB", (10, 10)))
    w = int(probe.textlength(label, font=FSM)) + 48
    img = Image.new("RGBA", (w, 56), (0, 0, 0, 0))
    d = ImageDraw.Draw(img)
    d.rounded_rectangle([0, 0, w - 1, 55], radius=28, fill=CHIP)
    d.text((24, 12), label, font=FSM, fill=SUBTLE)
    return img


def spreadsheet_card():
    from openpyxl import load_workbook
    wb = load_workbook(ROOT.parent / "exports" / "jam-scn-1-forecast.xlsx")
    ws = wb["Forecast"]
    body = [[c.value for c in r[:6]] for r in ws.iter_rows(min_row=1, max_row=10)]
    tail = [[c.value for c in r[:6]] for r in ws.iter_rows(min_row=ws.max_row - 2,
                                                           max_row=ws.max_row)]
    rows = body + tail
    img = Image.new("RGB", (W, H), "#F4F6F8")
    d = ImageDraw.Draw(img)
    d.rectangle([0, 0, W, 170], fill="#1D6F42")
    d.text((60, 55), "jam-scn-1-forecast.xlsx", font=FH, fill="white")
    d.text((60, 110), "Forecast · LineItems · Actuals", font=FSM, fill="#BFE3CF")
    fs = ImageFont.truetype(f"{DJ}/DejaVuSans.ttf", 30)
    fsb = ImageFont.truetype(f"{DJ}/DejaVuSans-Bold.ttf", 30)
    pick = [0, 1, 2, 5]
    col_w = [430, 250, 150, 190]
    y = 230
    for ri, row in enumerate(rows):
        x = 30
        header = ri == 0
        totals = ri >= len(rows) - 3
        for k, ci in enumerate(pick):
            val = row[ci] if ci < len(row) else None
            if header and k == 3:
                val = "Total"
            if header and k == 2:
                val = "Wk 1"
            cw = col_w[k]
            fill = "#DDE7EE" if header else ("#EAF3EC" if totals else
                                             ("white" if ri % 2 else "#EFF4F7"))
            d.rectangle([x, y, x + cw, y + 74], fill=fill, outline="#C9D4DC")
            txt = "" if val is None else (f"{val:,.0f}" if isinstance(val, (int, float)) else str(val))
            if len(txt) > 26:
                txt = txt[:25] + "…"
            font = fsb if (header or totals) else fs
            color = "#C0392B" if (isinstance(val, (int, float)) and val < 0) else "#203542"
            d.text((x + 14, y + 20), txt, font=font, fill=color)
            x += cw
        y += 74
    d.text((60, y + 44), "Every number traces to a line item.", font=FMED, fill="#5C7080")
    d.text((60, y + 116), "Grid · line items · actuals — 3 sheets.", font=FMED, fill="#5C7080")
    return img


# ------------------------------------------------ Allica brand cards
def logo_block(d, cx, cy, scale=1.0):
    """Cashew wordmark + 'from Allica Bank' byline, centered on (cx, cy)."""
    flogo = ImageFont.truetype(f"{DJ}/DejaVuSans-Bold.ttf", int(150 * scale))
    fby = ImageFont.truetype(f"{DJ}/DejaVuSans-Bold.ttf", int(54 * scale))
    wm = "Cashew"
    ww = d.textlength(wm, font=flogo)
    x0 = cx - ww / 2
    d.text((x0, cy - int(105 * scale)), wm, font=flogo, fill=NAVY)
    by = "from Allica Bank"
    bw_ = d.textlength(by, font=fby)
    d.text((x0 + ww - bw_, cy + int(75 * scale)), by, font=fby, fill=SLATE)


def intro_card(a=1.0):
    img = Image.new("RGB", (W, H), PAPER)
    d = ImageDraw.Draw(img)
    logo_block(d, W // 2, H // 2 - 120)
    tag = "Cashflow forecasting, in your chat"
    tw = d.textlength(tag, font=FMED)
    d.text(((W - tw) // 2, H // 2 + 190), tag, font=FMED, fill="#4A566E")
    if a < 1:
        img = Image.blend(Image.new("RGB", (W, H), PAPER), img, a)
    return img


def ending_card():
    img = Image.new("RGB", (W, H), PAPER)
    d = ImageDraw.Draw(img)
    logo_block(d, W // 2, H // 2 - 160)
    lines = [("Forecast. Export. Reconcile.", FMEDB, NAVY),
             ("Real Open Banking + Xero data", FMED, "#4A566E"),
             ("@cashewcashflowbot", FMEDB, "#2563EB")]
    y = H // 2 + 160
    for txt, font, col in lines:
        tw = d.textlength(txt, font=font)
        d.text(((W - tw) // 2, y), txt, font=font, fill=col)
        y += 86
    return img


ASKS = ["Who owes me money?", "Is my VAT covered?", "What's on direct debit?",
        "Set my cash floor at £50k", "What if revenue drops 30%?",
        "We're hiring at £4k/mo from Sept", "Compare June vs May",
        "Why did the forecast miss?"]

CAPS = [("check", "3-month forecast, verdict first"),
        ("check", "Reconciles vs actuals — says why"),
        ("chart", "One-ask Excel export"),
        ("check", "Debtor chase list & aging"),
        ("tri", "VAT pot tracking & accrual check"),
        ("check", "Direct debits & standing orders"),
        ("bulb", "Scenarios: late payers, hires, capex"),
        ("bulb", "Learns your business every month")]


def asks_frame(t, t0):
    img = Image.new("RGB", (W, H), NAVY)
    d = ImageDraw.Draw(img)
    title = "Just ask…"
    d.text(((W - d.textlength(title, font=FLOGO_S)) // 2, 170), title,
           font=FLOGO_S, fill="white")
    y = 330
    for i, q in enumerate(ASKS):
        at = t0 + 0.35 + i * 0.32
        if t < at:
            break
        prog = min(1.0, (t - at) / 0.22)
        e = prog * prog * (3 - 2 * prog)
        pw = int(d.textlength(q, font=FPILL)) + 90
        x = (W - pw) // 2
        yy = y + int((1 - e) * 26)
        a = int(255 * e)
        pill = Image.new("RGBA", (pw, 96), (0, 0, 0, 0))
        pd = ImageDraw.Draw(pill)
        side = i % 2 == 0
        pd.rounded_rectangle([0, 0, pw - 1, 95], radius=48,
                             fill=(43, 82, 120, a) if side else (24, 37, 51, a))
        pd.text((45, 24), q, font=FPILL, fill=(255, 255, 255, a))
        img.paste(pill, (x, yy), pill)
        y += 122
    foot = "…it answers with real numbers"
    d.text(((W - d.textlength(foot, font=FMED)) // 2, H - 300),
           foot, font=FMED, fill="#9FB3D9")
    return img


def caps_frame(t, t0):
    img = Image.new("RGB", (W, H), NAVY)
    d = ImageDraw.Draw(img)
    title = "Everything on board"
    d.text(((W - d.textlength(title, font=FLOGO_S)) // 2, 170), title,
           font=FLOGO_S, fill="white")
    y = 350
    for i, (icon, txt) in enumerate(CAPS):
        at = t0 + 0.3 + i * 0.28
        if t < at:
            break
        prog = min(1.0, (t - at) / 0.22)
        e = prog * prog * (3 - 2 * prog)
        a = int(255 * e)
        xx = 90 + int((1 - e) * 40)
        row = Image.new("RGBA", (W, 100), (0, 0, 0, 0))
        rd = ImageDraw.Draw(row)
        rd.ellipse([0, 22, 56, 78], fill=(43, 82, 120, a))
        draw_icon(rd, icon, (255, 255, 255, a), 8, 30)
        rd.text((90, 26), txt, font=FMED, fill=(255, 255, 255, a))
        img.paste(row, (xx, y), row)
        y += 128
    foot = "Deterministic engine — the AI never invents a number"
    d.text(((W - d.textlength(foot, font=FSM)) // 2, H - 240),
           foot, font=FSM, fill="#9FB3D9")
    return img


# ------------------------------------------------ chat scene
def chrome(canvas):
    canvas.rectangle([0, 0, W, 60], fill=HEADER)
    canvas.text((40, 14), "09:41", font=FSM, fill=TEXT)
    canvas.text((W - 150, 14), "5G  ▮▮▮", font=FSM, fill=TEXT)
    canvas.rectangle([0, 60, W, 200], fill=HEADER)
    canvas.text((44, 100), "←", font=FH, fill=ACCENT)
    canvas.ellipse([120, 82, 216, 178], fill=NAVY)
    canvas.text((148, 100), "C", font=ImageFont.truetype(
        f"{DJ}/DejaVuSans-Bold.ttf", 56), fill="white")
    canvas.text((246, 88), "Cashew Cashflow", font=FH, fill=TEXT)
    canvas.text((246, 140), "bot", font=FSM, fill=ACCENT)
    canvas.rectangle([0, H - 130, W, H], fill=HEADER)
    canvas.rounded_rectangle([100, H - 112, W - 130, H - 26], radius=40, fill="#242F3D")
    canvas.text((140, H - 92), "Message", font=F, fill=SUBTLE)
    canvas.text((36, H - 100), "+", font=FH, fill=SUBTLE)
    canvas.ellipse([W - 108, H - 112, W - 22, H - 26], fill=ACCENT)
    canvas.polygon([(W - 82, H - 88), (W - 44, H - 69), (W - 82, H - 50)], fill="white")


def typing_bubble(phase):
    img = Image.new("RGBA", (170, 90), (0, 0, 0, 0))
    d = ImageDraw.Draw(img)
    d.rounded_rectangle([0, 0, 169, 89], radius=26, fill=IN_BUBBLE)
    for i in range(3):
        dy = -10 * max(0.0, math.sin(phase * 2 * math.pi + i * 0.9))
        cx = 40 + i * 45
        d.ellipse([cx - 11, 48 + dy - 11, cx + 11, 48 + dy + 11], fill=SUBTLE)
    return img


def ease(t):
    return t * t * (3 - 2 * t)


def load(p):
    return (ROOT / "assets" / p).read_text().strip()


_ol = load("out_outlook.txt").split("\n")
_weeks_i = next(i for i, l in enumerate(_ol) if l.startswith("_Weeks"))
_vat_i = next(i for i, l in enumerate(_ol) if l.startswith("*VAT"))
outlook = "\n".join(_ol[:_weeks_i + 6] + ["  … and 9 more weeks"]
                    + [""] + _ol[_vat_i:_vat_i + 2])

recon_lines = load("out_reconcile.txt").split("\n")
keep = list(range(0, 5)) + [5, 8, 9, 11, 12]
keep += [i for i, l in enumerate(recon_lines) if "surprises" in l.lower()][:1]
keep += [i for i, l in enumerate(recon_lines) if "Director" in l][:1]
keep += [i for i, l in enumerate(recon_lines) if "_Lessons:_" in l][:1]
keep += [i for i, l in enumerate(recon_lines) if "weekly sweep" in l][:1]
recon = "\n".join(recon_lines[i] for i in sorted(set(keep)) if i < len(recon_lines))

MSGS = {
    "u1": make_bubble("Am I okay? Show me the next 3 months", True, "09:41"),
    "b1": make_bubble(outlook, False, "09:41"),
    "u2": make_bubble("Nice — export it to Excel 📊", True, "09:42"),
    "b2": make_doc_bubble("jam-scn-1-forecast.xlsx", "9.7 KB · xlsx",
                          "Forecast grid + line items + actuals.", "09:42"),
    "u3": make_bubble("Why did the forecast miss? Reconcile against actuals",
                      True, "17:20"),
    "b3": make_bubble(recon, False, "17:20"),
}
CHIP_JUN = make_chip("Tuesday · 30 June")
CHIP_JUL = make_chip("Friday · 31 July")

# ---- 2.5x-pace timeline ----------------------------------------------------
SCRIPT = [
    (2.5, "chip", "jun"),
    (2.7, "msg", "u1"), (3.1, "typing_on", None), (4.1, "typing_off", None),
    (4.1, "msg", "b1"),
    (8.9, "msg", "u2"), (9.3, "typing_on", None), (10.1, "typing_off", None),
    (10.1, "msg", "b2"),
    (15.3, "chip", "jul"),
    (15.5, "msg", "u3"), (15.9, "typing_on", None), (17.0, "typing_off", None),
    (17.0, "msg", "b3"),
]
CAPTIONS = [
    (4.4, 8.6, "① 3-month forecast, verdict first"),
    (10.3, 11.5, "② One-ask Excel export"),
    (17.4, 22.8, "③ Reconciles vs actuals — and says why"),
]
INTRO_END = 2.4
SHEET_T0, SHEET_T1 = 11.6, 15.2
CHAT_END = 23.2
ASKS_T0, ASKS_T1 = 23.2, 29.6
CAPS_T0, CAPS_T1 = 29.6, 35.8
END_T0 = 35.8
END_T = 40.5
XFADE = 0.45

SHEET = spreadsheet_card()
ENDING = ending_card()

GAP = 26
X_IN, X_OUT = 30, W - 30


def chat_items(t):
    items = []
    typing_since = None
    for at, kind, key in SCRIPT:
        if at > t:
            break
        if kind == "chip":
            img = CHIP_JUN if key == "jun" else CHIP_JUL
            items.append((img, (W - img.width) // 2, at))
        elif kind == "msg":
            img = MSGS[key]
            x = X_OUT - img.width if key.startswith("u") else X_IN
            items.append((img, x, at))
        elif kind == "typing_on":
            typing_since = at
        elif kind == "typing_off":
            typing_since = None
    if typing_since is not None:
        items.append((typing_bubble((t * 2.2) % 1.0), X_IN, typing_since))
    return items


def chat_frame(t):
    img = Image.new("RGB", (W, H), BG)
    d = ImageDraw.Draw(img)
    items = chat_items(t)
    y = H - 150
    placed = []
    for img_i, x, at in reversed(items):
        y -= img_i.height + GAP
        placed.append((img_i, x, y, at))
    for img_i, x, y, at in placed:
        if y + img_i.height < 200 or y > H - 130:
            continue
        prog = min(1.0, max(0.0, (t - at) / 0.18))
        e = ease(prog)
        if e < 1.0:
            s = 0.84 + 0.16 * e
            sw, sh = max(1, int(img_i.width * s)), max(1, int(img_i.height * s))
            small = img_i.resize((sw, sh))
            ghost = small.copy()
            ghost.putalpha(small.split()[3].point(lambda a_: int(a_ * e)))
            img.paste(ghost, (x + (img_i.width - sw) // 2,
                              y + (img_i.height - sh)), ghost)
        else:
            img.paste(img_i, (x, y), img_i)
    chrome(d)
    for t0, t1, txt in CAPTIONS:
        if t0 <= t <= t1:
            a = min(1.0, (t - t0) / 0.3, (t1 - t) / 0.3)
            pw = int(d.textlength(txt, font=FMED)) + 80
            pill = Image.new("RGBA", (pw, 92), (0, 0, 0, 0))
            pd = ImageDraw.Draw(pill)
            pd.rounded_rectangle([0, 0, pw - 1, 91], radius=46,
                                 fill=(20, 33, 46, int(235 * a)))
            pd.text((40, 22), txt, font=FMED, fill=(255, 255, 255, int(255 * a)))
            img.paste(pill, ((W - pw) // 2, 230), pill)
    return img


def _sheet_zoom(t):
    z = 1.0 + 0.06 * ease(min(1.0, max(0.0, (t - SHEET_T0) / (SHEET_T1 - SHEET_T0))))
    zw, zh = int(W * z), int(H * z)
    img = SHEET.resize((zw, zh))
    return img.crop(((zw - W) // 2, (zh - H) // 2,
                     (zw - W) // 2 + W, (zh - H) // 2 + H))


def blend(a_img, b_img, a):
    return Image.blend(a_img, b_img, ease(min(1.0, max(0.0, a))))


def frame(t):
    if t < INTRO_END:
        return intro_card(min(1.0, t / 0.6))
    if t < INTRO_END + XFADE:
        return blend(intro_card(), chat_frame(t), (t - INTRO_END) / XFADE)
    if SHEET_T0 - XFADE <= t < SHEET_T0:
        return blend(chat_frame(t), _sheet_zoom(SHEET_T0), (t - SHEET_T0 + XFADE) / XFADE)
    if SHEET_T0 <= t < SHEET_T1:
        return _sheet_zoom(t)
    if SHEET_T1 <= t < SHEET_T1 + XFADE:
        return blend(_sheet_zoom(SHEET_T1), chat_frame(t), (t - SHEET_T1) / XFADE)
    if t < CHAT_END:
        return chat_frame(t)
    if t < CHAT_END + XFADE:
        return blend(chat_frame(CHAT_END), asks_frame(t, ASKS_T0), (t - CHAT_END) / XFADE)
    if t < ASKS_T1:
        return asks_frame(t, ASKS_T0)
    if t < ASKS_T1 + XFADE:
        return blend(asks_frame(ASKS_T1, ASKS_T0), caps_frame(t, CAPS_T0),
                     (t - ASKS_T1) / XFADE)
    if t < END_T0:
        return caps_frame(t, CAPS_T0)
    if t < END_T0 + XFADE:
        return blend(caps_frame(END_T0, CAPS_T0), ENDING, (t - END_T0) / XFADE)
    return ENDING


def main():
    n = int(END_T * FPS)
    for i in range(n):
        frame(i / FPS).save(FRAMES / f"f{i:05d}.png")
        if i % 300 == 0:
            print(f"frame {i}/{n}")
    print(f"done: {n} frames")


if __name__ == "__main__":
    main()
