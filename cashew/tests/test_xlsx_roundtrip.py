"""Excel round-trip: export a workbook, edit the LineItems sheet with openpyxl,
plan_import -> apply_import, and verify store state / history / undo.

In-process only — a Store on a tmp sqlite path, no live server, and a minimal
synthetic compute-result dict instead of the real engine."""
import json

import pytest
from openpyxl import load_workbook

from engine.export import LINEITEMS_SHEET, export_xlsx
from engine.store import Store
from engine.xlsx_import import apply_import, plan_import

ORG = "test-org"

# smallest dict export_xlsx reads: buckets/occurrences for the Forecast grid,
# plus the three balance scalars (empty enriched list -> empty Actuals sheet)
MINIMAL_RESULT = {"buckets": [], "occurrences": [], "projected_net": 0.0,
                  "close": 0.0, "opening_balance": 1000.0}


@pytest.fixture
def store(tmp_path):
    s = Store(tmp_path / "engine-test.db")
    yield s
    s.close()


def seed(s: Store) -> dict[str, int]:
    return {
        "rent": s.upsert_item(ORG, "Rent — Landlord Ltd", "rent", "recurring_fixed",
                              {"amount": -1200.0, "day_of_month": 1}),
        "payroll": s.upsert_item(ORG, "Payroll — Staff", "payroll",
                                 "recurring_variable",
                                 {"observations": [-5000.0, -5100.0, -4950.0],
                                  "lookback_months": 3, "day_of_month": 25}),
        "saas": s.upsert_item(ORG, "SaaS — Toolco", "subscription_saas",
                              "recurring_fixed",
                              {"amount": -49.99, "day_of_month": 5}),
        "repair": s.upsert_item(ORG, "Kiln repair", "repairs_maintenance", "one_off",
                                {"amount": -900.0, "date": "2026-08-15"}),
    }


def export(s: Store, tmp_path, name="wb.xlsx") -> str:
    path = str(tmp_path / name)
    export_xlsx(path, MINIMAL_RESULT, s.items(ORG), [])
    return path


def sheet_map(ws):
    """-> (header_row_no, {header_lower: col_no}, {row_name: row_no})."""
    hrow = next(r for r in range(1, ws.max_row + 1)
                if str(ws.cell(row=r, column=1).value).strip() == "ID")
    cols = {str(c.value).strip().lower(): c.column for c in ws[hrow] if c.value}
    names = {ws.cell(row=r, column=cols["name"]).value: r
             for r in range(hrow + 1, ws.max_row + 1)
             if ws.cell(row=r, column=cols["name"]).value}
    return hrow, cols, names


# --- export shape ---------------------------------------------------------------

def test_workbook_structure_and_params_are_json(store, tmp_path):
    """Params must be json.dumps output (regression: str() wrote Python repr)."""
    ids = seed(store)
    wb = load_workbook(export(store, tmp_path))
    assert wb.sheetnames == ["Forecast", LINEITEMS_SHEET, "Actuals"]
    ws = wb[LINEITEMS_SHEET]
    hrow, cols, names = sheet_map(ws)
    header = [ws.cell(row=hrow, column=c).value
              for c in range(1, len(cols) + 1)]
    assert header == ["ID", "Name", "Category", "Method", "Params", "Start", "End",
                      "Source", "Locked", "Note"]
    assert header[0] == "ID"                       # stable id is the FIRST column
    # instruction banner above the header names the editable columns
    banner = " ".join(str(ws.cell(row=r, column=1).value) for r in range(1, hrow))
    assert "leave ID blank" in banner and "End" in banner
    by_id = {it["id"]: it for it in store.items(ORG)}
    assert len(names) == len(ids)
    for name, rno in names.items():
        raw = ws.cell(row=rno, column=cols["params"]).value
        parsed = json.loads(raw)                   # str(dict) repr would raise here
        assert parsed == by_id[ws.cell(row=rno, column=cols["id"]).value]["params"]


# --- the full round trip ----------------------------------------------------------

def test_roundtrip_plan_apply_history_undo(store, tmp_path):
    ids = seed(store)
    path = export(store, tmp_path)

    wb = load_workbook(path)
    ws = wb[LINEITEMS_SHEET]
    _, cols, names = sheet_map(ws)
    # 1) edit one amount inside Params (rent -1200 -> -1350.55)
    ws.cell(row=names["Rent — Landlord Ltd"], column=cols["params"],
            value=json.dumps({"amount": -1350.55, "day_of_month": 1}))
    # 2) change only the End date (saas) -> an ending, not a generic change
    ws.cell(row=names["SaaS — Toolco"], column=cols["end"], value="2026-09-30")
    # 3) delete one row (payroll) -> a removal proposal
    ws.delete_rows(names["Payroll — Staff"], 1)
    # 4) add one new row with a blank ID -> an addition
    new = [None] * len(cols)
    new[cols["name"] - 1] = "Retainer — New Consultant"
    new[cols["category"] - 1] = "consulting_fees"
    new[cols["method"] - 1] = "recurring_fixed"
    new[cols["params"] - 1] = json.dumps({"amount": -800.0, "day_of_month": 15})
    new[cols["start"] - 1] = "2026-08-01"
    new[cols["note"] - 1] = "added by owner in Excel"
    ws.append(new)
    wb.save(path)

    plan = plan_import(path, store, ORG)
    json.dumps(plan)                               # plan must be JSON-able
    assert plan["skipped"] == []
    assert [c["id"] for c in plan["changes"]] == [ids["rent"]]
    assert set(plan["changes"][0]["field_changes"]) == {"params"}
    assert plan["changes"][0]["field_changes"]["params"]["new"]["amount"] == -1350.55
    assert plan["endings"] == [{"id": ids["saas"], "name": "SaaS — Toolco",
                                "end_date": "2026-09-30", "note": ""}]
    assert [a["name"] for a in plan["additions"]] == ["Retainer — New Consultant"]
    assert plan["additions"][0]["method"] == "recurring_fixed"
    assert plan["additions"][0]["start_date"] == "2026-08-01"
    assert plan["removals"] == [{"id": ids["payroll"], "name": "Payroll — Staff"}]
    assert "1 change(s)" in plan["summary"] and "1 removal(s)" in plan["summary"]

    res = apply_import(plan, store, ORG)
    assert res["errors"] == []
    assert res["applied"] == {"changes": 1, "additions": 1, "endings": 1,
                              "removals": 1}

    # store state
    rent = store.get_item(ORG, ids["rent"])
    assert rent["params"]["amount"] == -1350.55
    assert rent["source"] == "owner" and rent["locked"] == 1
    saas = store.get_item(ORG, ids["saas"])
    assert saas["end_date"] == "2026-09-30" and saas["locked"] == 1
    added = store.get_item(ORG, "Retainer — New Consultant")
    assert added and added["source"] == "owner" and added["locked"] == 1
    assert added["params"] == {"amount": -800.0, "day_of_month": 15}
    assert store.get_item(ORG, ids["payroll"])["active"] == 0
    assert all(it["name"] != "Payroll — Staff" for it in store.items(ORG))

    # history rows exist for each applied write (after the 4 seed creates)
    actions = [r["action"] for r in store.conn.execute(
        "SELECT action FROM line_item_history WHERE org=? ORDER BY id", (ORG,))]
    assert actions[:4] == ["create"] * 4
    assert actions[4:] == ["update", "create", "end", "deactivate"]

    # undo reverts the last write (payroll deactivation)
    msg = store.undo_last(ORG)
    assert "Payroll" in msg
    assert store.get_item(ORG, ids["payroll"])["active"] == 1


# --- malformed rows land in skipped, never raise -----------------------------------

def test_malformed_params_json_is_skipped(store, tmp_path):
    ids = seed(store)
    path = export(store, tmp_path)
    wb = load_workbook(path)
    ws = wb[LINEITEMS_SHEET]
    _, cols, names = sheet_map(ws)
    ws.cell(row=names["Rent — Landlord Ltd"], column=cols["params"],
            value="{'amount': -1300.0,")            # repr-ish, not JSON
    wb.save(path)

    plan = plan_import(path, store, ORG)
    assert plan["changes"] == [] and plan["endings"] == [] and plan["additions"] == []
    assert len(plan["skipped"]) == 1
    assert "JSON" in plan["skipped"][0]["reason"]
    # the row is present in the sheet, so it must NOT be proposed for removal
    assert ids["rent"] not in [r["id"] for r in plan["removals"]]
    assert plan["removals"] == []


def test_unknown_category_row_is_skipped(store, tmp_path):
    seed(store)
    path = export(store, tmp_path)
    wb = load_workbook(path)
    ws = wb[LINEITEMS_SHEET]
    _, cols, _ = sheet_map(ws)
    new = [None] * len(cols)
    new[cols["name"] - 1] = "Jetpack lease"
    new[cols["category"] - 1] = "flying_cars"       # not in taxonomy registry
    new[cols["method"] - 1] = "recurring_fixed"
    new[cols["params"] - 1] = json.dumps({"amount": -500.0, "day_of_month": 1})
    ws.append(new)
    wb.save(path)

    plan = plan_import(path, store, ORG)
    assert plan["additions"] == []
    assert len(plan["skipped"]) == 1
    assert "flying_cars" in plan["skipped"][0]["reason"]


def test_bad_method_and_bad_date_are_skipped(store, tmp_path):
    ids = seed(store)
    path = export(store, tmp_path)
    wb = load_workbook(path)
    ws = wb[LINEITEMS_SHEET]
    _, cols, names = sheet_map(ws)
    ws.cell(row=names["Kiln repair"], column=cols["method"], value="sometimes")
    ws.cell(row=names["SaaS — Toolco"], column=cols["end"], value="next month")
    wb.save(path)

    plan = plan_import(path, store, ORG)
    reasons = " | ".join(s["reason"] for s in plan["skipped"])
    assert len(plan["skipped"]) == 2
    assert "sometimes" in reasons and "next month" in reasons
    assert plan["changes"] == [] and plan["endings"] == []
    assert plan["removals"] == []                   # both rows still in the sheet
    assert ids  # seeded ok


def test_untouched_workbook_is_a_no_op_plan(store, tmp_path):
    seed(store)
    plan = plan_import(export(store, tmp_path), store, ORG)
    assert plan["changes"] == [] and plan["additions"] == []
    assert plan["endings"] == [] and plan["removals"] == []
    assert plan["skipped"] == []
    res = apply_import(plan, store, ORG)
    assert res == {"applied": {"changes": 0, "additions": 0, "endings": 0,
                               "removals": 0}, "errors": []}
