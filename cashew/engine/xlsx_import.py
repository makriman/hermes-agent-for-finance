"""Excel round-trip import — the other half of the PRD's 'augment Excel' bridge.

The owner edits the LineItems sheet of an exported workbook; `plan_import`
diffs it against the store (by the stable ID column) into a JSON-able plan,
and `apply_import` applies that plan through the existing store API only
(upsert_item / end_item / deactivate_item), so line_item_history and undo
keep working.  Pure library: no CLI, no prints, and malformed rows land in
plan['skipped'] with a reason instead of raising.

Forecast and Actuals sheets are ignored entirely.
"""
from __future__ import annotations

import json
from datetime import date, datetime

from openpyxl import load_workbook

from . import taxonomy as tax
from .export import LINEITEMS_SHEET

VALID_METHODS = ("recurring_fixed", "recurring_variable", "one_off", "linked")

# store fields the owner may edit through the sheet (Source/Locked/ID are not)
EDITABLE_FIELDS = ("name", "category", "method", "params", "start_date", "end_date")

# sheet header text (lowercased) -> store field
_HEADER_MAP = {
    "id": "id", "name": "name", "category": "category", "method": "method",
    "params": "params", "start": "start_date", "start_date": "start_date",
    "end": "end_date", "end_date": "end_date", "note": "note",
}


# --- cell / value normalisation ------------------------------------------------

def _norm(v):
    """Excel cell value -> plain Python: '' -> None, dates -> ISO, floats rounded."""
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    if isinstance(v, bool):
        return v
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, float):
        return int(v) if v.is_integer() else round(v, 2)
    return v


def _round_floats(v):
    """Round floats (2dp) recursively — sensible for money-shaped params."""
    if isinstance(v, bool):
        return v
    if isinstance(v, float):
        return round(v, 2)
    if isinstance(v, list):
        return [_round_floats(x) for x in v]
    if isinstance(v, dict):
        return {k: _round_floats(x) for k, x in v.items()}
    return v


def _canon(v):
    """Canonical form for equality tests: 5 == 5.0 == 5.001-rounding noise."""
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    if isinstance(v, list):
        return [_canon(x) for x in v]
    if isinstance(v, dict):
        return {k: _canon(x) for k, x in sorted(v.items())}
    return v


def _parse_date(v, field: str):
    """-> (iso_or_none, error_or_none). Empty cells are legitimate (open window)."""
    if v is None:
        return None, None
    try:
        return date.fromisoformat(str(v)[:10]).isoformat(), None
    except ValueError:
        return None, f"{field} '{v}' is not a valid YYYY-MM-DD date"


def _parse_params(v):
    """-> (dict_or_none, error_or_none)."""
    if v is None:
        return None, "Params is empty — expected a JSON object like {\"amount\": -100}"
    if not isinstance(v, str):
        return None, f"Params must be JSON text, got {type(v).__name__} '{v}'"
    try:
        p = json.loads(v)
    except json.JSONDecodeError as e:
        return None, f"Params is not valid JSON ({e.msg} at position {e.pos})"
    if not isinstance(p, dict):
        return None, "Params JSON must be an object like {\"amount\": -100}"
    return _round_floats(p), None


def _find_header(ws):
    """Locate the header row (has both ID and Name) -> (row_no, {field: col_no})."""
    for r in range(1, min(ws.max_row, 10) + 1):
        lowered = [str(_norm(c.value) or "").lower() for c in ws[r]]
        if "id" in lowered and "name" in lowered:
            cols = {}
            for idx, h in enumerate(lowered, start=1):
                field = _HEADER_MAP.get(h)
                if field and field not in cols:
                    cols[field] = idx
            return r, cols
    return None, {}


# --- plan --------------------------------------------------------------------

def plan_import(path: str, store, org: str) -> dict:
    """Diff the workbook's LineItems sheet against store.items(org) by ID.

    Returns a JSON-able plan:
      {changes:   [{id, name, field_changes: {field: {old, new}}, note}],
       additions: [{name, category, method, params, start_date, end_date, note}],
       endings:   [{id, name, end_date, note}],
       removals:  [{id, name}],
       skipped:   [{row, reason}],
       summary:   str}

    Never raises on bad content — malformed rows land in `skipped`.
    """
    plan = {"changes": [], "additions": [], "endings": [], "removals": [],
            "skipped": [], "summary": ""}

    def skip(row: int, reason: str):
        plan["skipped"].append({"row": row, "reason": reason})

    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:  # unreadable / not an xlsx
        skip(0, f"could not open workbook: {e}")
        plan["summary"] = "Nothing to import: workbook could not be read."
        return plan
    if LINEITEMS_SHEET not in wb.sheetnames:
        skip(0, f"workbook has no '{LINEITEMS_SHEET}' sheet")
        plan["summary"] = f"Nothing to import: no {LINEITEMS_SHEET} sheet."
        return plan
    ws = wb[LINEITEMS_SHEET]
    hrow, cols = _find_header(ws)
    if hrow is None:
        skip(0, f"no header row with ID and Name found on {LINEITEMS_SHEET}")
        plan["summary"] = "Nothing to import: LineItems header row not found."
        return plan

    items = store.items(org)                       # active items only
    by_id = {it["id"]: it for it in items}
    by_name = {it["name"]: it for it in items}
    seen_ids: set[int] = set()
    seen_new_names: set[str] = set()

    for rno in range(hrow + 1, ws.max_row + 1):
        row = {f: _norm(ws.cell(row=rno, column=c).value) for f, c in cols.items()}
        if all(v is None for v in row.values()):
            continue                                # blank spacer row
        note = row.get("note") or ""

        # ---- rows carrying an ID: change / ending / unchanged --------------
        if row.get("id") is not None:
            try:
                iid = int(row["id"])
            except (TypeError, ValueError):
                skip(rno, f"ID '{row['id']}' is not a number")
                continue
            if iid in seen_ids:
                skip(rno, f"duplicate row for ID {iid}")
                continue
            it = by_id.get(iid)
            if it is None:
                skip(rno, f"ID {iid} does not match any active line item")
                continue
            seen_ids.add(iid)                       # present in sheet -> not a removal

            errs, new = [], {}
            new["name"] = row.get("name")
            if new["name"] is None:
                errs.append("Name is blank")
            new["category"] = row.get("category")
            if new["category"] is None:
                errs.append("Category is blank")
            elif new["category"] != it["category"] and new["category"] not in tax._C:
                errs.append(f"unknown category '{new['category']}'")
            new["method"] = row.get("method")
            if new["method"] not in VALID_METHODS:
                errs.append(f"method '{new['method']}' is not one of {list(VALID_METHODS)}")
            new["params"], perr = _parse_params(row.get("params"))
            if perr:
                errs.append(perr)
            new["start_date"], derr = _parse_date(row.get("start_date"), "Start")
            if derr:
                errs.append(derr)
            new["end_date"], derr = _parse_date(row.get("end_date"), "End")
            if derr:
                errs.append(derr)
            if errs:
                skip(rno, "; ".join(errs))
                continue
            if (new["name"] != it["name"] and new["name"] in by_name
                    and by_name[new["name"]]["id"] != iid):
                skip(rno, f"rename to '{new['name']}' collides with existing "
                          f"item id {by_name[new['name']]['id']}")
                continue

            field_changes = {}
            for f in EDITABLE_FIELDS:
                old, nv = it[f], new[f]
                if f == "params":
                    if _canon(old) != _canon(nv):
                        field_changes[f] = {"old": old, "new": nv}
                elif (old or None) != (nv or None):
                    field_changes[f] = {"old": old, "new": nv}
            if not field_changes:
                continue                            # untouched row
            if set(field_changes) == {"end_date"} and field_changes["end_date"]["new"]:
                plan["endings"].append({"id": iid, "name": it["name"],
                                        "end_date": field_changes["end_date"]["new"],
                                        "note": note})
            else:
                plan["changes"].append({"id": iid, "name": it["name"],
                                        "field_changes": field_changes, "note": note})
            continue

        # ---- rows with a blank ID: additions --------------------------------
        name = row.get("name")
        if name is None:
            skip(rno, "row has content but no Name (and no ID)")
            continue
        if name in by_name:
            skip(rno, f"Name '{name}' already exists (id {by_name[name]['id']}) — "
                      "to edit it, keep its ID in the row")
            continue
        if name in seen_new_names:
            skip(rno, f"duplicate new row for Name '{name}'")
            continue
        errs = []
        cat = row.get("category")
        if cat is None:
            errs.append("Category is blank")
        elif cat not in tax._C:
            errs.append(f"unknown category '{cat}'")
        meth = row.get("method")
        if meth not in VALID_METHODS:
            errs.append(f"method '{meth}' is not one of {list(VALID_METHODS)}")
        params, perr = _parse_params(row.get("params"))
        if perr:
            errs.append(perr)
        start, derr = _parse_date(row.get("start_date"), "Start")
        if derr:
            errs.append(derr)
        end, derr = _parse_date(row.get("end_date"), "End")
        if derr:
            errs.append(derr)
        if errs:
            skip(rno, "; ".join(errs))
            continue
        seen_new_names.add(name)
        plan["additions"].append({"name": name, "category": cat, "method": meth,
                                  "params": params, "start_date": start,
                                  "end_date": end, "note": note})

    # ---- active store items missing from the sheet: propose deactivation ----
    for it in items:
        if it["id"] not in seen_ids:
            plan["removals"].append({"id": it["id"], "name": it["name"]})

    plan["summary"] = (
        f"{len(plan['changes'])} change(s), {len(plan['additions'])} addition(s), "
        f"{len(plan['endings'])} ending(s), {len(plan['removals'])} removal(s) "
        f"(deactivate, never delete); {len(plan['skipped'])} row(s) skipped.")
    return plan


# --- apply -------------------------------------------------------------------

def apply_import(plan: dict, store, org: str) -> dict:
    """Apply a plan from plan_import through the store API only.

    Every write goes via upsert_item(source='owner', locked=1), end_item or
    deactivate_item, so line_item_history is written and store.undo_last works.
    Returns {applied: {changes, additions, endings, removals}, errors: [...]}.
    """
    applied = {"changes": 0, "additions": 0, "endings": 0, "removals": 0}
    errors: list[str] = []

    for ch in plan.get("changes", []):
        try:
            it = store.get_item(org, int(ch["id"]))
            if it is None or not it.get("active", 1):
                errors.append(f"change: item id {ch['id']} is no longer active")
                continue
            fc = ch.get("field_changes", {})
            new = {f: (fc[f]["new"] if f in fc else it[f]) for f in EDITABLE_FIELDS}
            note = ch.get("note") or "edited via xlsx import"
            if new["name"] != it["name"]:
                # store is keyed by (org, name): a rename is create-under-new-name
                # + deactivate-old — both audited, and undo reverts the deactivate
                store.upsert_item(org, new["name"], new["category"], new["method"],
                                  new["params"], counterparty=it.get("counterparty") or "",
                                  start_date=new["start_date"], end_date=new["end_date"],
                                  source="owner", locked=1,
                                  note=f"{note} (renamed from '{it['name']}')")
                store.deactivate_item(org, it["id"],
                                      note=f"renamed to '{new['name']}' via xlsx import")
            else:
                store.upsert_item(org, it["name"], new["category"], new["method"],
                                  new["params"], counterparty=it.get("counterparty") or "",
                                  start_date=new["start_date"], end_date=new["end_date"],
                                  source="owner", locked=1, note=note)
            applied["changes"] += 1
        except Exception as e:
            errors.append(f"change id {ch.get('id')}: {e}")

    for add in plan.get("additions", []):
        try:
            store.upsert_item(org, add["name"], add["category"], add["method"],
                              add.get("params") or {}, counterparty="",
                              start_date=add.get("start_date"),
                              end_date=add.get("end_date"),
                              source="owner", locked=1,
                              note=add.get("note") or "added via xlsx import")
            applied["additions"] += 1
        except Exception as e:
            errors.append(f"addition '{add.get('name')}': {e}")

    for e_ in plan.get("endings", []):
        try:
            if store.end_item(org, int(e_["id"]), e_["end_date"],
                              note=e_.get("note") or "ended via xlsx import"):
                applied["endings"] += 1
            else:
                errors.append(f"ending: item id {e_['id']} not found")
        except Exception as e:
            errors.append(f"ending id {e_.get('id')}: {e}")

    for rm in plan.get("removals", []):
        try:
            if store.deactivate_item(org, int(rm["id"]),
                                     note="removed from workbook via xlsx import"):
                applied["removals"] += 1
            else:
                errors.append(f"removal: item id {rm['id']} not found")
        except Exception as e:
            errors.append(f"removal id {rm.get('id')}: {e}")

    return {"applied": applied, "errors": errors}
